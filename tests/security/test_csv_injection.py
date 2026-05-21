"""Defend against CSV-injection (a.k.a. formula injection) in report outputs.

When a downstream user opens ``quality_report.xlsx`` in Excel, any cell that
begins with ``=``, ``+``, ``-``, ``@``, tab, or carriage-return is parsed as
a formula by default.  An attacker who can write the underlying data file
can stage commands like ``=cmd|'/c calc'!A1`` and have them execute on the
analyst's machine.

The validators' ``message`` field is the most likely vector — it
interpolates user-controlled column values.  This test runs the pipeline
against a fixture whose ``species_common`` contains formula-style payloads
and asserts that no cell in the generated XLSX is a live formula.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from data_quality_monitor.pipeline import run_pipeline

pytestmark = pytest.mark.security

REPO_ROOT = Path(__file__).resolve().parents[2]
RULES_FILE = REPO_ROOT / "tests" / "fixtures" / "sparrowhawk" / "sparrowhawk_rules.yml"


HEADER = (
    "ringing_scheme,band_token,band_id,species_common,permit_code,bander_license,"
    "flyway_region,last_resight_date,migration_status,lifecycle_state,habitat_class,"
    "capture_method,morph_variant,taxon_code,first_banded_date,band_status,"
    "active_resights,observation_notes\n"
)

ROWS = [
    'Euring,Euring:RA-1,RA-1,"=cmd|\'/c calc\'!A1",EU-PERM-1,BL-1,WEU,,Never resighted,Alive,WET,MIST,,,,Active,,',
    'NABU,NABU:RA-2,RA-2,"+SUM(1+1)",EU-PERM-2,BL-2,EAA,,Never resighted,Alive,TER,MIST,,,,Issued,,',
    'Euring,Euring:RA-3,RA-3,"-2+3+cmd",EU-PERM-3,BL-3,NAM,,Never resighted,Alive,WET,MIST,,,,Active,,',
    'Euring,Euring:RA-4,RA-4,"@SUM(A1:A10)",EU-PERM-4,BL-4,EAF,,Never resighted,Alive,WET,MIST,,,,Active,,',
]


def test_xlsx_report_has_no_live_formulas(tmp_path: Path) -> None:
    fixture = tmp_path / "injected.csv"
    fixture.write_text(HEADER + "\n".join(ROWS) + "\n", encoding="utf-8")

    run_pipeline(input_path=fixture, rules_path=RULES_FILE, output_dir=tmp_path)
    xlsx_path = tmp_path / "quality_report.xlsx"
    assert xlsx_path.exists()

    workbook = load_workbook(xlsx_path)
    dangerous_prefixes = ("=", "+", "-", "@", "\t", "\r")
    offenders: list[tuple[str, str, str]] = []
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith(dangerous_prefixes):
                    # openpyxl marks live formulas with data_type == "f".
                    if cell.data_type == "f":
                        offenders.append((sheet, cell.coordinate, value))
    assert not offenders, (
        "Live formulas leaked into the report:\n"
        + "\n".join(f"  {s}!{c}: {v!r}" for s, c, v in offenders)
    )
