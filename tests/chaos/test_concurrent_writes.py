"""Two pipeline runs writing to the *same* output_dir at the same time
must not corrupt each other's artefacts.

The current implementation in ``write_reports`` does not use temp-file +
rename, so this test documents the *observed* behaviour: artefacts after
both runs are valid (parseable JSON / CSV / XLSX), even if the content is
whichever run wrote last.  If we later move to atomic writes, this test
becomes a stronger guarantee.
"""

from __future__ import annotations

import json
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from data_quality_monitor.pipeline import run_pipeline

pytestmark = pytest.mark.chaos

REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURE = REPO_ROOT / "tests" / "fixtures" / "sparrowhawk" / "rings_control.csv"
RULES = REPO_ROOT / "tests" / "fixtures" / "sparrowhawk" / "sparrowhawk_rules.yml"


def _go(output_dir: Path) -> None:
    run_pipeline(input_path=FIXTURE, rules_path=RULES, output_dir=output_dir)


def test_two_concurrent_runs_leave_valid_artefacts(tmp_path: Path) -> None:
    with ThreadPoolExecutor(max_workers=2) as pool:
        futures = [pool.submit(_go, tmp_path) for _ in range(2)]
        for f in futures:
            f.result()  # surface any exception

    # All three artefacts must exist and parse cleanly.
    summary = json.loads((tmp_path / "summary.json").read_text(encoding="utf-8"))
    assert "quality_score" in summary

    report = pd.read_csv(tmp_path / "quality_report.csv")
    assert not report.empty

    wb = load_workbook(tmp_path / "quality_report.xlsx")
    assert {"summary", "checks"}.issubset(set(wb.sheetnames))
